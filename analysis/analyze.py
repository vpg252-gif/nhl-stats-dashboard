"""
NHL Analysis Engine
===================
Reads from the SQLite database and produces:
  1. Terminal report — printed summary of key findings
  2. Excel workbook  — multi-sheet formatted export

Analyses included:
  - Skater rankings (points, goals, assists, shooting %)
  - Team performance (goals for/against, PP%, PK%, goal diff)
  - Advanced stats (points per game, goals per game, goal contributions)

Usage:
    pip install openpyxl pandas
    python analysis/analyze.py
"""

import sys
import os
import sqlite3
import logging
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR    = Path(__file__).resolve().parent.parent
DB_PATH     = BASE_DIR / "data" / "nhl_stats.db"
OUTPUT_DIR  = BASE_DIR / "outputs"
EXCEL_PATH  = OUTPUT_DIR / "nhl_analysis.xlsx"

OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ---------------------------------------------------------------------------
# Colours (openpyxl uses ARGB hex, no #)
# ---------------------------------------------------------------------------
NHL_DARK    = "001F5B"   # deep navy
NHL_LIGHT   = "C8102E"   # NHL red
HEADER_BG   = "001F5B"   # navy header background
HEADER_FG   = "FFFFFF"   # white header text
ALT_ROW     = "EEF2FF"   # light blue-grey alternating row
GOLD        = "FFD700"   # top-3 highlight
SILVER      = "C0C0C0"
BRONZE      = "CD7F32"
WHITE       = "FFFFFF"
DARK_TEXT   = "1A1A2E"


# ---------------------------------------------------------------------------
# Database
# ---------------------------------------------------------------------------

def get_conn() -> sqlite3.Connection:
    if not DB_PATH.exists():
        raise FileNotFoundError(
            f"Database not found at {DB_PATH}. "
            "Run 'python db/load_database.py' first."
        )
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def query(conn, sql: str, params=()) -> pd.DataFrame:
    return pd.read_sql_query(sql, conn, params=params)


# ---------------------------------------------------------------------------
# Analysis functions — return DataFrames
# ---------------------------------------------------------------------------

def skater_rankings(conn) -> pd.DataFrame:
    """Top skaters ranked by points with supporting stats."""
    sql = """
        SELECT
            full_name                                   AS "Player",
            team_abbrev                                 AS "Team",
            position                                    AS "Pos",
            COALESCE(goals, 0)                          AS "Goals",
            COALESCE(assists, 0)                        AS "Assists",
            COALESCE(points, 0)                         AS "Points",
            COALESCE(plus_minus, 0)                     AS "+/-",
            COALESCE(shots, 0)                          AS "Shots",
            COALESCE(hits, 0)                           AS "Hits",
            COALESCE(blocked_shots, 0)                  AS "Blocks",
            COALESCE(pp_goals, 0)                       AS "PP Goals",
            COALESCE(gw_goals, 0)                       AS "GWG",
            ROUND(COALESCE(shooting_pct, 0), 1)         AS "Sh%",
            ROUND(COALESCE(points_per_game, 0), 3)      AS "P/GP"
        FROM skater_stats
        WHERE points IS NOT NULL
        ORDER BY points DESC, goals DESC
        LIMIT 100
    """
    return query(conn, sql)


def goal_scorers(conn) -> pd.DataFrame:
    """Top goal scorers with shooting efficiency."""
    sql = """
        SELECT
            full_name                               AS "Player",
            team_abbrev                             AS "Team",
            position                                AS "Pos",
            COALESCE(goals, 0)                      AS "Goals",
            COALESCE(shots, 0)                      AS "Shots",
            ROUND(COALESCE(shooting_pct, 0), 1)     AS "Sh%",
            COALESCE(pp_goals, 0)                   AS "PP Goals",
            COALESCE(sh_goals, 0)                   AS "SH Goals",
            COALESCE(gw_goals, 0)                   AS "GWG",
            ROUND(COALESCE(goals_per_game, 0), 3)   AS "G/GP"
        FROM skater_stats
        WHERE goals IS NOT NULL AND goals > 0
        ORDER BY goals DESC, shooting_pct DESC
        LIMIT 50
    """
    return query(conn, sql)


def assist_leaders(conn) -> pd.DataFrame:
    """Top playmakers ranked by assists."""
    sql = """
        SELECT
            full_name                               AS "Player",
            team_abbrev                             AS "Team",
            position                                AS "Pos",
            COALESCE(assists, 0)                    AS "Assists",
            COALESCE(goals, 0)                      AS "Goals",
            COALESCE(points, 0)                     AS "Points",
            ROUND(
                CASE WHEN COALESCE(points,0) > 0
                THEN CAST(COALESCE(assists,0) AS REAL) / points * 100
                ELSE 0 END, 1
            )                                       AS "Ast%"
        FROM skater_stats
        WHERE assists IS NOT NULL AND assists > 0
        ORDER BY assists DESC
        LIMIT 50
    """
    return query(conn, sql)


def advanced_stats(conn) -> pd.DataFrame:
    """
    Advanced skater metrics:
      - Points per game
      - Goals per game
      - Goal contribution % (what share of points are goals vs assists)
      - Physical presence score (hits + blocks)
    """
    sql = """
        SELECT
            full_name                                           AS "Player",
            team_abbrev                                         AS "Team",
            position                                            AS "Pos",
            COALESCE(points, 0)                                 AS "Points",
            ROUND(COALESCE(points_per_game, 0), 3)              AS "P/GP",
            ROUND(COALESCE(goals_per_game, 0), 3)               AS "G/GP",
            ROUND(
                CASE WHEN COALESCE(points,0) > 0
                THEN CAST(COALESCE(goals,0) AS REAL) / points * 100
                ELSE 0 END, 1
            )                                                   AS "Goal Contrib%",
            ROUND(
                CASE WHEN COALESCE(points,0) > 0
                THEN CAST(COALESCE(assists,0) AS REAL) / points * 100
                ELSE 0 END, 1
            )                                                   AS "Ast Contrib%",
            COALESCE(plus_minus, 0)                             AS "+/-",
            COALESCE(hits, 0) + COALESCE(blocked_shots, 0)     AS "Physical Score",
            COALESCE(penalty_minutes, 0)                        AS "PIM",
            ROUND(COALESCE(shooting_pct, 0), 1)                 AS "Sh%"
        FROM skater_stats
        WHERE points IS NOT NULL AND points >= 10
        ORDER BY points_per_game DESC
        LIMIT 75
    """
    return query(conn, sql)


def team_performance(conn) -> pd.DataFrame:
    """Team-level offensive, defensive, and special teams stats."""
    sql = """
        SELECT
            team_name                                               AS "Team",
            team_abbrev                                             AS "Abbrev",
            conference                                              AS "Conference",
            division                                                AS "Division",
            wins                                                    AS "W",
            losses                                                  AS "L",
            ot_losses                                               AS "OTL",
            points                                                  AS "PTS",
            games_played                                            AS "GP",
            goals_for                                               AS "GF",
            goals_against                                           AS "GA",
            goal_diff                                               AS "GDiff",
            ROUND(CAST(goals_for AS REAL) / NULLIF(games_played,0), 2)      AS "GF/GP",
            ROUND(CAST(goals_against AS REAL) / NULLIF(games_played,0), 2)  AS "GA/GP",
            ROUND(pp_pct, 1)                                        AS "PP%",
            ROUND(pk_pct, 1)                                        AS "PK%",
            home_wins                                               AS "Home W",
            home_losses                                             AS "Home L",
            away_wins                                               AS "Away W",
            away_losses                                             AS "Away L",
            l10_wins                                                AS "L10 W",
            l10_losses                                              AS "L10 L"
        FROM standings
        ORDER BY points DESC, wins DESC
    """
    return query(conn, sql)


def goalie_rankings(conn) -> pd.DataFrame:
    """Goalie leaderboard by save percentage."""
    sql = """
        SELECT
            full_name                           AS "Goalie",
            team_abbrev                         AS "Team",
            COALESCE(wins, 0)                   AS "Wins",
            ROUND(COALESCE(save_pct, 0), 3)     AS "SV%",
            ROUND(COALESCE(gaa, 0), 2)          AS "GAA",
            COALESCE(shutouts, 0)               AS "SO"
        FROM goalie_stats
        WHERE wins IS NOT NULL AND wins > 0
        ORDER BY save_pct DESC, wins DESC
        LIMIT 50
    """
    return query(conn, sql)


def position_breakdown(conn) -> pd.DataFrame:
    """Average stats broken down by position."""
    sql = """
        SELECT
            position                                        AS "Position",
            COUNT(*)                                        AS "Players",
            ROUND(AVG(COALESCE(goals,0)), 1)                AS "Avg Goals",
            ROUND(AVG(COALESCE(assists,0)), 1)              AS "Avg Assists",
            ROUND(AVG(COALESCE(points,0)), 1)               AS "Avg Points",
            ROUND(AVG(COALESCE(shooting_pct,0)), 1)         AS "Avg Sh%",
            ROUND(AVG(COALESCE(hits,0)), 1)                 AS "Avg Hits",
            ROUND(AVG(COALESCE(blocked_shots,0)), 1)        AS "Avg Blocks",
            ROUND(AVG(COALESCE(plus_minus,0)), 1)           AS "Avg +/-"
        FROM skater_stats
        WHERE position IN ('C','L','R','D')
          AND points IS NOT NULL
        GROUP BY position
        ORDER BY FIELD(position, 'C','L','R','D')
    """
    # FIELD() is MySQL syntax — use CASE for SQLite
    sql = """
        SELECT
            position                                        AS "Position",
            COUNT(*)                                        AS "Players",
            ROUND(AVG(COALESCE(goals,0)), 1)                AS "Avg Goals",
            ROUND(AVG(COALESCE(assists,0)), 1)              AS "Avg Assists",
            ROUND(AVG(COALESCE(points,0)), 1)               AS "Avg Points",
            ROUND(AVG(COALESCE(shooting_pct,0)), 1)         AS "Avg Sh%",
            ROUND(AVG(COALESCE(hits,0)), 1)                 AS "Avg Hits",
            ROUND(AVG(COALESCE(blocked_shots,0)), 1)        AS "Avg Blocks",
            ROUND(AVG(COALESCE(plus_minus,0)), 1)           AS "Avg +/-"
        FROM skater_stats
        WHERE position IN ('C','L','R','D')
          AND points IS NOT NULL
        GROUP BY position
        ORDER BY
            CASE position
                WHEN 'C' THEN 1
                WHEN 'L' THEN 2
                WHEN 'R' THEN 3
                WHEN 'D' THEN 4
            END
    """
    return query(conn, sql)


# ---------------------------------------------------------------------------
# Terminal report
# ---------------------------------------------------------------------------

def print_section(title: str):
    print("\n" + "="*60)
    print(f"  {title}")
    print("="*60)


def print_terminal_report(analyses: dict):
    print("\n" + "█"*60)
    print("  NHL STATS ANALYSIS REPORT")
    print("█"*60)

    # Points leaders
    print_section("TOP 20 POINTS LEADERS")
    df = analyses["skater_rankings"].head(20)
    print(f"  {'#':<4} {'Player':<25} {'Team':<6} {'Pos':<4} {'G':<5} {'A':<5} {'PTS':<5} {'P/GP':<6} {'Sh%'}")
    print("  " + "-"*58)
    for i, row in df.iterrows():
        rank = i + 1
        medal = "🥇" if rank == 1 else "🥈" if rank == 2 else "🥉" if rank == 3 else f"{rank}. "
        print(f"  {str(medal):<4} {row['Player']:<25} {row['Team']:<6} {row['Pos']:<4} "
              f"{row['Goals']:<5} {row['Assists']:<5} {row['Points']:<5} "
              f"{row['P/GP']:<6} {row['Sh%']}")

    # Goal scorers
    print_section("TOP 10 GOAL SCORERS")
    df = analyses["goal_scorers"].head(10)
    print(f"  {'#':<4} {'Player':<25} {'Team':<6} {'Goals':<7} {'Shots':<7} {'Sh%':<7} {'PPG':<6} {'GWG'}")
    print("  " + "-"*58)
    for i, row in df.iterrows():
        print(f"  {i+1:<4} {row['Player']:<25} {row['Team']:<6} {row['Goals']:<7} "
              f"{row['Shots']:<7} {row['Sh%']:<7} {row['PP Goals']:<6} {row['GWG']}")

    # Team performance
    print_section("TEAM STANDINGS & PERFORMANCE")
    df = analyses["team_performance"].head(32)
    print(f"  {'#':<4} {'Team':<25} {'Div':<10} {'W':<4} {'L':<4} {'OTL':<5} {'PTS':<5} {'GF/GP':<7} {'GA/GP':<7} {'PP%':<6} {'PK%'}")
    print("  " + "-"*75)
    for i, row in df.iterrows():
        print(f"  {i+1:<4} {row['Team']:<25} {row['Division']:<10} {row['W']:<4} {row['L']:<4} "
              f"{row['OTL']:<5} {row['PTS']:<5} {row['GF/GP']:<7} {row['GA/GP']:<7} "
              f"{row['PP%']:<6} {row['PK%']}")

    # Advanced stats
    print_section("TOP 10 — ADVANCED STATS (min. 10 pts, ranked by P/GP)")
    df = analyses["advanced_stats"].head(10)
    print(f"  {'#':<4} {'Player':<25} {'Team':<6} {'PTS':<5} {'P/GP':<6} {'G/GP':<6} {'Goal%':<8} {'Phys':<6} {'+/-'}")
    print("  " + "-"*65)
    for i, row in df.iterrows():
        print(f"  {i+1:<4} {row['Player']:<25} {row['Team']:<6} {row['Points']:<5} "
              f"{row['P/GP']:<6} {row['G/GP']:<6} {row['Goal Contrib%']:<8} "
              f"{row['Physical Score']:<6} {row['+/-']}")

    # Goalie leaders
    print_section("TOP 10 GOALIES (by SV%)")
    df = analyses["goalie_rankings"].head(10)
    print(f"  {'#':<4} {'Goalie':<25} {'Team':<6} {'W':<5} {'SV%':<7} {'GAA':<6} {'SO'}")
    print("  " + "-"*50)
    for i, row in df.iterrows():
        print(f"  {i+1:<4} {row['Goalie']:<25} {row['Team']:<6} {row['Wins']:<5} "
              f"{row['SV%']:<7} {row['GAA']:<6} {row['SO']}")

    # Position breakdown
    print_section("AVERAGE STATS BY POSITION")
    df = analyses["position_breakdown"]
    print(f"  {'Pos':<6} {'Players':<9} {'Avg G':<8} {'Avg A':<8} {'Avg PTS':<10} {'Avg Sh%':<9} {'Avg Hits':<10} {'Avg Blk':<9} {'Avg +/-'}")
    print("  " + "-"*70)
    for _, row in df.iterrows():
        print(f"  {row['Position']:<6} {row['Players']:<9} {row['Avg Goals']:<8} "
              f"{row['Avg Assists']:<8} {row['Avg Points']:<10} {row['Avg Sh%']:<9} "
              f"{row['Avg Hits']:<10} {row['Avg Blocks']:<9} {row['Avg +/-']}")

    print("\n" + "█"*60)
    print("  END OF REPORT")
    print("█"*60 + "\n")


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def make_header_style():
    return {
        "font":      Font(name="Arial", bold=True, color=HEADER_FG, size=10),
        "fill":      PatternFill("solid", start_color=HEADER_BG),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border":    Border(
            bottom=Side(style="medium", color=HEADER_FG),
        ),
    }


def make_title_style():
    return {
        "font":      Font(name="Arial", bold=True, color=WHITE, size=14),
        "fill":      PatternFill("solid", start_color=NHL_LIGHT),
        "alignment": Alignment(horizontal="left", vertical="center"),
    }


def apply_style(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)


def write_sheet(
    wb: Workbook,
    sheet_name: str,
    title: str,
    df: pd.DataFrame,
    medal_col: str = None,
    number_formats: dict = None,
):
    """
    Write a DataFrame to an Excel sheet with full formatting.

    Parameters
    ----------
    medal_col     : Column name to apply gold/silver/bronze highlighting to top 3 rows.
    number_formats: Dict of {col_name: excel_format_string} e.g. {"Sh%": "0.0"}
    """
    ws = wb.create_sheet(sheet_name)
    number_formats = number_formats or {}

    # --- Title row ---
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell = ws.cell(row=1, column=1, value=title)
    apply_style(title_cell, make_title_style())
    ws.row_dimensions[1].height = 28

    # --- Header row ---
    header_style = make_header_style()
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=2, column=col_idx, value=col_name)
        apply_style(cell, header_style)
    ws.row_dimensions[2].height = 22

    # --- Data rows ---
    medal_colors = {1: GOLD, 2: "C8C8C8", 3: "CD7F32"}

    for row_idx, (_, row) in enumerate(df.iterrows(), 3):
        data_row_num = row_idx - 2  # 1-based data row
        is_alt = data_row_num % 2 == 0
        row_bg = ALT_ROW if is_alt else WHITE

        # Medal highlight for top 3
        if medal_col and data_row_num in medal_colors:
            row_bg = medal_colors[data_row_num]

        for col_idx, col_name in enumerate(df.columns, 1):
            value = row[col_name]
            # Convert numpy types to native python
            if hasattr(value, "item"):
                value = value.item()

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name="Arial", size=9, color=DARK_TEXT)
            cell.fill = PatternFill("solid", start_color=row_bg)
            cell.alignment = Alignment(
                horizontal="center" if col_idx > 1 else "left",
                vertical="center"
            )
            cell.border = Border(
                bottom=Side(style="hair", color="CCCCCC")
            )

            # Number formatting
            if col_name in number_formats:
                cell.number_format = number_formats[col_name]

        ws.row_dimensions[row_idx].height = 16

    # --- Column widths ---
    for col_idx, col_name in enumerate(df.columns, 1):
        col_letter = get_column_letter(col_idx)
        # Measure max content width
        max_len = max(
            len(str(col_name)),
            df[col_name].astype(str).str.len().max() if len(df) > 0 else 0
        )
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 8), 30)

    # Freeze header rows
    ws.freeze_panes = "A3"

    return ws


def build_summary_sheet(wb: Workbook, analyses: dict):
    """Create a summary/dashboard sheet as the first tab."""
    ws = wb.create_sheet("📊 Summary", 0)
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "🏒  NHL STATS ANALYSIS DASHBOARD"
    title.font = Font(name="Arial", bold=True, size=16, color=WHITE)
    title.fill = PatternFill("solid", start_color=NHL_DARK)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # Subtitle
    ws.merge_cells("A2:H2")
    sub = ws["A2"]
    sub.value = "Generated by NHL Stats Project  |  Source: NHL Stats API"
    sub.font = Font(name="Arial", size=10, color="666666", italic=True)
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Section helper
    def section_header(row, col, text):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = Font(name="Arial", bold=True, size=10, color=WHITE)
        cell.fill = PatternFill("solid", start_color=NHL_LIGHT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 20
        return cell

    def data_cell(row, col, value, bold=False, num_format=None):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = Font(name="Arial", size=9, bold=bold, color=DARK_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if num_format:
            cell.number_format = num_format
        ws.row_dimensions[row].height = 15
        return cell

    # --- Points leaders ---
    section_header(4, 1, "🏆  TOP 10 POINTS LEADERS")
    ws.merge_cells("A4:D4")
    headers = ["Player", "Team", "G", "PTS"]
    for i, h in enumerate(headers):
        c = ws.cell(row=5, column=i+1, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=HEADER_FG)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    skaters = analyses["skater_rankings"].head(10)
    for r_idx, (_, row) in enumerate(skaters.iterrows(), 6):
        bg = [GOLD, "C8C8C8", "CD7F32"][r_idx - 6] if r_idx <= 8 else (ALT_ROW if r_idx % 2 == 0 else WHITE)
        for c_idx, val in enumerate([row["Player"], row["Team"], row["Goals"], row["Points"]], 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val if not hasattr(val, "item") else val.item())
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

    # --- Goal leaders ---
    section_header(4, 6, "🎯  TOP 10 GOAL SCORERS")
    ws.merge_cells("F4:I4")
    headers2 = ["Player", "Team", "Goals", "Sh%"]
    for i, h in enumerate(headers2):
        c = ws.cell(row=5, column=i+6, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=HEADER_FG)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    goals_df = analyses["goal_scorers"].head(10)
    for r_idx, (_, row) in enumerate(goals_df.iterrows(), 6):
        bg = [GOLD, "C8C8C8", "CD7F32"][r_idx - 6] if r_idx <= 8 else (ALT_ROW if r_idx % 2 == 0 else WHITE)
        for c_idx, val in enumerate([row["Player"], row["Team"], row["Goals"], row["Sh%"]], 6):
            c = ws.cell(row=r_idx, column=c_idx, value=val if not hasattr(val, "item") else val.item())
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center" if c_idx > 6 else "left")

    # --- Team leaders ---
    row_start = 18
    section_header(row_start, 1, "🏒  TOP 10 TEAMS BY POINTS")
    ws.merge_cells(f"A{row_start}:G{row_start}")
    t_headers = ["Team", "Div", "W", "L", "OTL", "PTS", "GDiff"]
    for i, h in enumerate(t_headers):
        c = ws.cell(row=row_start+1, column=i+1, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=HEADER_FG)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    teams_df = analyses["team_performance"].head(10)
    for r_idx, (_, row) in enumerate(teams_df.iterrows(), row_start+2):
        bg = ALT_ROW if r_idx % 2 == 0 else WHITE
        vals = [row["Team"], row["Division"], row["W"], row["L"], row["OTL"], row["PTS"], row["GDiff"]]
        for c_idx, val in enumerate(vals, 1):
            c = ws.cell(row=r_idx, column=c_idx, value=val if not hasattr(val, "item") else val.item())
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center" if c_idx > 1 else "left")

    # --- Goalie leaders ---
    section_header(row_start, 9, "🧤  TOP 5 GOALIES (SV%)")
    ws.merge_cells(f"I{row_start}:L{row_start}")
    g_headers = ["Goalie", "Team", "W", "SV%"]
    for i, h in enumerate(g_headers):
        c = ws.cell(row=row_start+1, column=i+9, value=h)
        c.font = Font(name="Arial", bold=True, size=9, color=HEADER_FG)
        c.fill = PatternFill("solid", start_color=HEADER_BG)
        c.alignment = Alignment(horizontal="center")

    goalies_df = analyses["goalie_rankings"].head(5)
    for r_idx, (_, row) in enumerate(goalies_df.iterrows(), row_start+2):
        bg = [GOLD, "C8C8C8", "CD7F32"][r_idx - (row_start+2)] if r_idx <= row_start+4 else (ALT_ROW if r_idx % 2 == 0 else WHITE)
        vals = [row["Goalie"], row["Team"], row["Wins"], row["SV%"]]
        for c_idx, val in enumerate(vals, 9):
            c = ws.cell(row=r_idx, column=c_idx, value=val if not hasattr(val, "item") else val.item())
            c.font = Font(name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=bg)
            c.alignment = Alignment(horizontal="center" if c_idx > 9 else "left")

    # Column widths
    col_widths = {
        "A": 26, "B": 8, "C": 6, "D": 6,
        "E": 6,  "F": 26, "G": 8, "H": 6,
        "I": 26, "J": 8, "K": 6, "L": 6,
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    ws.sheet_view.showGridLines = False


def export_excel(analyses: dict) -> Path:
    """Build and save the full Excel workbook."""
    logger.info("── Building Excel workbook...")
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Summary dashboard
    build_summary_sheet(wb, analyses)

    # Detailed sheets
    write_sheet(
        wb, "🏒 Skater Rankings", "NHL Skater Rankings — Points Leaders",
        analyses["skater_rankings"],
        medal_col="Points",
        number_formats={"Sh%": "0.0", "P/GP": "0.000"},
    )
    write_sheet(
        wb, "🎯 Goal Scorers", "NHL Goal Scorers",
        analyses["goal_scorers"],
        medal_col="Goals",
        number_formats={"Sh%": "0.0", "G/GP": "0.000"},
    )
    write_sheet(
        wb, "🍎 Assist Leaders", "NHL Assist Leaders — Playmakers",
        analyses["assist_leaders"],
        medal_col="Assists",
        number_formats={"Ast%": "0.0"},
    )
    write_sheet(
        wb, "📈 Advanced Stats", "NHL Advanced Stats (min. 10 pts, ranked by P/GP)",
        analyses["advanced_stats"],
        medal_col="P/GP",
        number_formats={"P/GP": "0.000", "G/GP": "0.000", "Goal Contrib%": "0.0", "Sh%": "0.0"},
    )
    write_sheet(
        wb, "🏒 Team Performance", "NHL Team Performance & Standings",
        analyses["team_performance"],
        medal_col="PTS",
        number_formats={"GF/GP": "0.00", "GA/GP": "0.00", "PP%": "0.0", "PK%": "0.0"},
    )
    write_sheet(
        wb, "🧤 Goalies", "NHL Goalie Rankings",
        analyses["goalie_rankings"],
        medal_col="SV%",
        number_formats={"SV%": "0.000", "GAA": "0.00"},
    )
    write_sheet(
        wb, "📊 Position Avg", "Average Stats by Position",
        analyses["position_breakdown"],
    )

    wb.save(EXCEL_PATH)
    logger.info(f"   Excel saved to: {EXCEL_PATH}")
    return EXCEL_PATH


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    logger.info("="*60)
    logger.info("NHL ANALYSIS ENGINE — Starting")
    logger.info("="*60)

    conn = get_conn()

    logger.info("── Running analyses...")
    analyses = {
        "skater_rankings":   skater_rankings(conn),
        "goal_scorers":      goal_scorers(conn),
        "assist_leaders":    assist_leaders(conn),
        "advanced_stats":    advanced_stats(conn),
        "team_performance":  team_performance(conn),
        "goalie_rankings":   goalie_rankings(conn),
        "position_breakdown":position_breakdown(conn),
    }
    conn.close()

    for name, df in analyses.items():
        logger.info(f"   {name}: {len(df)} rows")

    # Terminal report
    print_terminal_report(analyses)

    # Excel export
    export_excel(analyses)

    logger.info("✅ Analysis complete!")
    logger.info(f"   Excel workbook: {EXCEL_PATH}")


if __name__ == "__main__":
    main()
